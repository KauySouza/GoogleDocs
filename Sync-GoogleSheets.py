# VMware PSO - v0.1
#
# This script syncs data between .csv and Google Sheets. That makes the Residency Dashboard on SharePoint more accurate and easy to maintain. The Google API version when this script was develop is 4.0.
#
#
# You will need:
# 	Internet access to connect on Google Sheet
#	A Google Account to authenticate.
#	Python 2.6 or later - Comes with your MAC.
#	A python package manager called "pip" - Usually comes with python 2.7 or later. Keep in mind that some bugs were found with El Captain Release. Try updating your python (brew install python).
#	Download the packages for API connection in your python repository - sudo pip install --upgrade google-api-python-client
#	Enable API connection on your google account - Follow only the Step 1 of this link https://developers.google.com/sheets/quickstart/python.
#   Install openpyxl python package - pip install openpyxl
#


#
# Importing packages and methods.
#
from __future__ import print_function
from apiclient import discovery
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage
import openpyxl
import datetime
import httplib2
import os


# Autheticating and storing credencials for further use.

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/sheets.googleapis.com-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/spreadsheets'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'DashboardResidencia'


def get_credentials():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'sheets.googleapis.com-python-dashboardresidencia.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

# Openning file.
filePath = "/Users/ksouza/Downloads/Relatorio_Quinzenal_v9.xlsx"
excelFile = open(filePath)


#Create main method
def main():
    #Setting up authentication
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?'
                    'version=v4')
    service = discovery.build('sheets', 'v4', http=http,
                              discoveryServiceUrl=discoveryUrl)
    
    #Identifing sheet - Those info are extrated from the sheet's URL'
    spreeadsheetId = '1xhsoObyX4DFvzLvx6rR6nD_Fie4ZxRFOocvHMQVBCA8'
    sheetId = '1160464292'

    #XLSX reading
    if filePath[-5:] == '.xlsx':
        wb = openpyxl.load_workbook(filename = filePath)
        tableSheet = wb['Atividades']

        #Dictionary to translate a simple index in A1 Notation
        column = {0 : 'A',
                1 : 'B',
                2 : 'C',
                3 : 'D',
                4 : 'E',
                5 : 'F',
                6 : 'G',
                7 : 'H',
                8 : 'I',
                9 : 'J',
                10 : 'K',
                11 : 'L',
                12 : 'M',
                13 : 'N',
                14 : 'O',
                15 : 'P',
                16 : 'Q',
            }
        
        #Defining Range to be written
        range_name = "Atividades!A1:" + str(column[tableSheet.max_column]) + str(tableSheet.max_row)
        #openpyxl.worksheet.Worksheet()

        #Building the content of the input
        for idxRow in xrange(6,tableSheet.max_row):
            row = []
            for idxCol in xrange(3,tableSheet.max_column):
                if tableSheet.cell(row = idxRow, column = idxCol).is_date:
                    row.append(str(tableSheet.cell(row = idxRow, column = idxCol).value))
                else:    
                    row.append(tableSheet.cell(row = idxRow, column = idxCol).value)
            inputList.append(row)
        
        print(inputList)
        
        
        body = {
            'values': inputList
        }
        result = service.spreadsheets().values().update(
        spreadsheetId=spreeadsheetId, range=range_name,
        valueInputOption="USER_ENTERED", body=body).execute()

        print("O resultado foi: " + str(result))

    else:
        print('The extention of the file is not suitable for this script')

#Running main method
if __name__ == '__main__':
    main()