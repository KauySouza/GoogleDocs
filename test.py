import openpyxl
import datetime
import httplib2

#Setting up authentication
filePath = "/Users/ksouza/Documents/Customer/Bradesco/Relatorio_Quinzenal_v6_ForMAC.xlsx"
#Identifing sheet
spreeadsheetId = '1xhsoObyX4DFvzLvx6rR6nD_Fie4ZxRFOocvHMQVBCA8'
sheetId = '1160464292'

#Example
#PUT https://sheets.googleapis.com/v4/spreadsheets/spreadsheetId/values/Sheet1!A1:D5?valueInputOption=USER_ENTERED

#XLSX reading
if filePath[-5:] == '.xlsx':
    wb = openpyxl.load_workbook(filename = filePath)
    tableSheet = wb['Atividades']

    #Dictionary to translate index in A1 Notation
    column = {0 : 'A',
            1 : 'B',
            2 : 'C',
            3 : 'D',
            4 : 'E',
            5 : 'F',
            6 : 'G',
            7 : 'H',
            8 : 'I',
            9 : 'J',
            10 : 'K',
            11 : 'L',
            12 : 'M',
            13 : 'N',
            14 : 'O',
            15 : 'P',
            16 : 'Q',
        }
    
    #Defining Range to be written
    range_name = "Sheet1!A1:" + str(column[tableSheet.max_column]) + str(tableSheet.max_row)

    inputList = []

    for idxRow in xrange(6,tableSheet.max_row):
        #print ("-"*40)
        row = []
        for idxCol in xrange(3,tableSheet.max_column):
            #print (tableSheet.cell(row = idxRow, column = idxCol).value)
            row.append(tableSheet.cell(row = idxRow, column = idxCol).value)
            
            
        inputList.append(row)

    
    #print(inputList[1:3])
    
    contentOfValues = ""
    for idx in xrange (0,len(inputList)):
        contentOfValues += str(inputList[idx])
    
    values = [
    [
        str(contentOfValues)
    ],
# Additional rows ...
    ]
    print (values)